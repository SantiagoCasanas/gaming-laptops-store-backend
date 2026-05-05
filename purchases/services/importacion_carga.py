"""
Importaciones — bulk update of OrdenCompra logistic state and costo_importacion
from an Excel file (template-based upload).

Workflow:
  1. User downloads the template (build_template_workbook).
  2. User fills it with rows from the byorderbox email — one row per package.
  3. User uploads it. parse_and_validate matches each row to an existing
     OrdenCompra by `numero_tracking` and either applies the changes (commit)
     or simulates them (dry-run).

Match policy: `numero_tracking` is the unique key. If a tracking number doesn't
exist in the system the row is reported as "no_mapeado" with the description
column so the user can identify what came in unexpected.

When matched and committed:
  • OrdenCompra.estado_logistico → 'en_oficina_importadora'
  • OrdenCompra.costo_importacion → valor en COP from the row (stored in COP per
    explicit user requirement, even though costo_compra is in USD)
  • OrdenCompra.fecha_en_oficina_importadora is auto-stamped by the model save()
  • UnidadProducto (if exists) → estado_producto='en_oficina_importadora' to
    mirror the logistic transition
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from purchases.models import OrdenCompra


# Plantilla — columnas (orden visible en el Excel)
TEMPLATE_COLUMNS = [
    {
        "key": "numero_tracking",
        "label": "Numero de tracking",
        "tipo": "texto",
        "required": True,
        "hint": "Track del paquete (clave de matching)",
    },
    {
        "key": "descripcion",
        "label": "Descripcion",
        "tipo": "texto",
        "required": False,
        "hint": "Solo informativa — ayuda a identificar items no mapeados",
    },
    {
        "key": "valor_importacion_cop",
        "label": "Valor importacion (COP)",
        "tipo": "numero",
        "required": True,
        "hint": "Valor por paquete en pesos colombianos",
    },
]


def build_template_workbook() -> Workbook:
    """
    Generate the Excel template for the importaciones upload. Row 1 = header,
    row 2 = type/required hint, data starts at row 3.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Importaciones"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2979C8")
    hint_font = Font(italic=True, color="6B7280", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, col in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

        hint = f"({col['tipo']}{' · obligatorio' if col['required'] else ''}) {col['hint']}"
        hint_cell = ws.cell(row=2, column=idx, value=hint)
        hint_cell.font = hint_font
        hint_cell.alignment = center

        ws.column_dimensions[cell.column_letter].width = max(22, len(col["label"]) + 6)

        # Force the tracking column to text so Excel doesn't convert long
        # numbers to scientific notation or strip leading zeros. The other
        # columns keep general formatting.
        if col["key"] == "numero_tracking":
            for row in range(3, 1003):
                ws.cell(row=row, column=idx).number_format = "@"

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"
    return wb


def _parse_decimal(raw):
    """
    Tolerant numeric parser for COP values. Recognized inputs:
      127435            → 127435
      127.435           → 127435   (es-CO thousand separator: dot + 3 digits)
      127,435           → 127435   (lone comma + 3 digits = thousand separator)
      "$ 127.435"       → 127435
      1.234.567         → 1234567
      1.234.567,89      → 1234567.89
      127,89            → 127.89   (lone comma + 1-2 digits = decimal)
      127.89            → 127.89   (lone dot + 1-2 digits = decimal)

    The catch is the COP convention: "127.435" looks like a Decimal 127.435 but
    is actually a thousand separator. Heuristic: a SINGLE dot/comma followed by
    EXACTLY 3 digits is treated as a thousand separator (COP prices are whole
    pesos, no decimals).
    """
    if raw is None:
        return None, None
    if isinstance(raw, (int, float, Decimal)):
        return Decimal(str(raw)), None
    s = str(raw).strip().replace("$", "").replace(" ", "")
    if s == "":
        return None, None

    # Both . and , present → es-CO with decimals (dots = thousands, comma = decimal)
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        # Multiple dots → all thousand separators
        s = s.replace(".", "")
    elif s.count(",") > 1:
        s = s.replace(",", "")
    elif "," in s:
        # Single comma. If followed by exactly 3 digits → thousand sep; else decimal.
        before, after = s.split(",")
        if len(after) == 3 and after.isdigit():
            s = before + after
        else:
            s = s.replace(",", ".")
    elif "." in s:
        # Single dot. Same heuristic — 3 digits after = es-CO thousand separator.
        before, after = s.split(".")
        if len(after) == 3 and after.isdigit():
            s = before + after
        # else leave as-is (decimal)

    try:
        return Decimal(s), None
    except (InvalidOperation, ValueError):
        return None, f'valor "{raw}" no es un numero valido'


def _normalize_track(raw):
    """
    Normalize a tracking value coming from Excel. openpyxl loads numeric cells
    as floats, so a tracking like 380534958280 arrives as 380534958280.0 — that
    breaks the equality match against the stored string "380534958280". We
    strip the trailing ".0" when the float has no fractional part. Whitespace
    inside the value is also removed (some carriers split the number with a
    space). Scientific notation (e.g. 9.63209e+18) is rendered as a fixed
    integer when possible.
    """
    if raw is None:
        return ""
    # Numeric types: int, float, Decimal — convert to integer string when there
    # is no fractional part, since tracking numbers are always whole.
    if isinstance(raw, bool):
        return ""
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        if raw.is_integer():
            # Use Decimal to avoid scientific notation for very large floats
            return format(int(raw), "d")
        return format(raw, "f").rstrip("0").rstrip(".")
    if isinstance(raw, Decimal):
        if raw == raw.to_integral_value():
            return format(raw.to_integral_value(), "f")
        return format(raw, "f").rstrip("0").rstrip(".")
    # Strings: trim and drop inner whitespace; also drop a trailing ".0" that
    # might come from Excel exports that already serialized the number.
    s = "".join(str(raw).split())
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _serialize_row(numero_tracking, descripcion, valor):
    """Build the `data` dict returned with every row in the preview."""
    return {
        "numero_tracking": numero_tracking or "",
        "descripcion": descripcion or "",
        "valor_importacion_cop": "" if valor is None else str(valor),
    }


def _apply_match(orden: OrdenCompra, valor_cop: Decimal, usuario):
    """
    Persist the match on the order. estado_logistico → en_oficina_importadora,
    costo_importacion → valor_cop. Also sync the linked UnidadProducto when it
    exists and is still a placeholder (no real serial yet).
    """
    with transaction.atomic():
        orden.estado_logistico = "en_oficina_importadora"
        orden.costo_importacion = valor_cop
        orden.usuario_ultima_modificacion = usuario
        orden.save()

        unidad = orden.unidad_producto
        if unidad and unidad.estado_producto == "viajando":
            unidad.estado_producto = "en_oficina_importadora"
            unidad.usuario_ultima_modificacion = usuario
            unidad.save(update_fields=["estado_producto", "usuario_ultima_modificacion"])


def parse_and_validate(uploaded_file, dry_run: bool, usuario) -> dict:
    """
    Parse the uploaded xlsx and return a preview dict similar to the bulk
    products upload:
        {
          "dry_run": bool,
          "total": int,
          "matched":   [{fila, data, orden_id, numero_orden, producto_nombre,
                         costo_anterior, costo_nuevo}],
          "no_mapeado": [{fila, data, motivo}],
          "fallidos":   [{fila, data, errores: [str]}],
        }
    """
    try:
        wb = load_workbook(BytesIO(uploaded_file.read()), data_only=True)
    except Exception as exc:
        return {
            "dry_run": dry_run,
            "total": 0,
            "matched": [],
            "no_mapeado": [],
            "fallidos": [
                {"fila": 0, "data": {}, "errores": [f"No se pudo leer el archivo: {exc}"]}
            ],
        }

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"dry_run": dry_run, "total": 0, "matched": [], "no_mapeado": [], "fallidos": []}

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]

    # Header → column index. Match by canonical key (lowercased, stripped).
    header_lookup = {h.lower(): i for i, h in enumerate(headers) if h}

    def col_index(*labels):
        # Accept multiple aliases for the same column. None = not found.
        for label in labels:
            for candidate in (label, label.replace("_", " ")):
                idx = header_lookup.get(candidate.lower())
                if idx is not None:
                    return idx
        return None

    track_col = col_index("Numero de tracking", "numero_tracking", "tracking", "track")
    desc_col = col_index("Descripcion", "descripcion", "item")
    valor_col = col_index(
        "Valor importacion (COP)",
        "valor_importacion_cop",
        "Valor importacion COP",
        "valor",
    )

    if track_col is None or valor_col is None:
        return {
            "dry_run": dry_run,
            "total": 0,
            "matched": [],
            "no_mapeado": [],
            "fallidos": [
                {
                    "fila": 0,
                    "data": {},
                    "errores": [
                        "El archivo no tiene las columnas requeridas: 'Numero de tracking' y 'Valor importacion (COP)'."
                    ],
                }
            ],
        }

    # Skip the hint row if present (row 2 of the template starts with '(')
    data_start = 1
    if len(rows) > 1:
        second_row = rows[1]
        if second_row and any(
            isinstance(v, str) and v.strip().startswith("(")
            for v in second_row
            if v is not None
        ):
            data_start = 2

    matched, no_mapeado, fallidos = [], [], []
    seen_tracks: set[str] = set()

    for row_idx, raw_row in enumerate(rows[data_start:], start=data_start + 1):
        if raw_row is None or all(v in (None, "") for v in raw_row):
            continue

        def cell(i):
            if i is None or i >= len(raw_row):
                return None
            return raw_row[i]

        numero_tracking = _normalize_track(cell(track_col))
        descripcion = (str(cell(desc_col) or "").strip()) if desc_col is not None else ""
        valor_raw = cell(valor_col)
        valor_cop, valor_err = _parse_decimal(valor_raw)

        data = _serialize_row(numero_tracking, descripcion, valor_cop)

        errores = []
        if not numero_tracking:
            errores.append("numero_tracking es obligatorio")
        if valor_cop is None and not valor_err:
            errores.append("valor_importacion_cop es obligatorio")
        if valor_err:
            errores.append(f"valor_importacion_cop: {valor_err}")

        if errores:
            fallidos.append({"fila": row_idx, "data": data, "errores": errores})
            continue

        if numero_tracking in seen_tracks:
            fallidos.append({
                "fila": row_idx,
                "data": data,
                "errores": [f"tracking '{numero_tracking}' aparece duplicado en el archivo"],
            })
            continue
        seen_tracks.add(numero_tracking)

        # Match against OrdenCompra.numero_tracking
        orden = OrdenCompra.objects.filter(numero_tracking=numero_tracking).first()

        if not orden:
            no_mapeado.append({
                "fila": row_idx,
                "data": data,
                "motivo": "no_se_encontro_orden_con_ese_tracking",
            })
            continue

        costo_anterior = orden.costo_importacion
        producto_nombre = orden.producto.nombre if orden.producto else "—"

        if not dry_run:
            try:
                _apply_match(orden, valor_cop, usuario)
            except Exception as exc:
                fallidos.append({
                    "fila": row_idx,
                    "data": data,
                    "errores": [f"error al guardar: {exc}"],
                })
                continue

        matched.append({
            "fila": row_idx,
            "data": data,
            "orden_id": orden.id,
            "numero_orden": orden.numero_orden,
            "producto_nombre": producto_nombre,
            "costo_anterior": str(costo_anterior) if costo_anterior is not None else None,
            "costo_nuevo": str(valor_cop),
        })

    return {
        "dry_run": dry_run,
        "total": len(matched) + len(no_mapeado) + len(fallidos),
        "matched": matched,
        "no_mapeado": no_mapeado,
        "fallidos": fallidos,
    }
