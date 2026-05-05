"""
Seed laptops inventory from Inventario_Laptops_Patecnologicos.xlsx.

Creates / upserts:
  • Brand (HP, ACER, ASUS, MSI, LENOVO, APPLE)
  • TipoProducto "Laptop" + 11 CampoProducto definitions
  • Proveedor default ("Inventario Local") for stock without supplier
  • Producto (one per Excel row) with ProductoCampoValor for every dynamic field
  • OrdenCompra (one per Excel row) which auto-creates the UnidadProducto.
    fecha_compra is distributed across March/April/May 2026 round-robin so the
    dashboard month filter has data on each of those months.
    The unit price is overridden after creation to match the Excel COP price.

Idempotent: re-running the command updates existing rows by stable keys.
"""
from datetime import date
from decimal import Decimal

from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from openpyxl import load_workbook

from products.models import (
    Brand, TipoProducto, CampoProducto, TipoProductoCampo,
    Proveedor, Producto, ProductoCampoValor, UnidadProducto,
)
from purchases.models import OrdenCompra

User = get_user_model()

EXCEL_PATH = 'products/datos-precarga/laptops/Inventario_Laptops_Patecnologicos.xlsx'

# Hardcoded TipoProducto definition. The 11 fields below match the Excel
# columns for laptops one-to-one.
LAPTOP_FIELDS = [
    'Modelo',
    'Procesador',
    'Tarjeta de Video',
    'Memoria',
    'Almacenamiento',
    'Tamaño',
    'Resolución de la pantalla',
    'Tasa de refresco',
    'Tipo de pantalla',
    'Conectividad',
    'Sistema Operativo',
]

# Excel "Estado" → UnidadProducto.condicion / OrdenCompra.condicion
ESTADO_TO_CONDICION = {
    'open box': 'open_box',
    'refurbished': 'refurbished',
    'reacondicionado': 'refurbished',
    'usada': 'usado',
    'usado': 'usado',
    'nueva': 'nuevo',
    'nuevo': 'nuevo',
}

# Excel "Entrega" → OrdenCompra.estado_logistico
ENTREGA_TO_ESTADO_LOGISTICO = {
    'stock': 'en_oficina',
    'viajando': 'viajando',
    'en_oficina_importadora': 'en_oficina_importadora',
}

# fecha_compra distribution: round-robin march/april/may 2026
FECHAS_COMPRA = [
    date(2026, 3, 5),
    date(2026, 3, 18),
    date(2026, 4, 3),
    date(2026, 4, 12),
    date(2026, 4, 22),
    date(2026, 5, 2),
    date(2026, 5, 14),
    date(2026, 5, 24),
]

# Default TRM if core.TRMHistory is empty — used to back-calc costo_compra (USD)
# from the Excel COP selling price. Only a placeholder for cost; the real
# selling price is overridden afterwards.
DEFAULT_TRM_COP = Decimal('3637.51')
ESTIMATED_MARGIN = Decimal('1.5')   # selling = 1.5 * cost in COP (rough)

PROVEEDOR_DEFAULT = 'Inventario Local'


class Command(BaseCommand):
    help = 'Carga inventario de laptops desde Inventario_Laptops_Patecnologicos.xlsx'

    def handle(self, *args, **options):
        usuario = self._get_superuser()
        if not usuario:
            return

        self.stdout.write(self.style.SUCCESS('\n=== CARGA DE LAPTOPS ==='))

        try:
            wb = load_workbook(EXCEL_PATH, data_only=True)
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'Archivo no encontrado: {EXCEL_PATH}'))
            return

        ws = wb['Inventario Laptops']
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[0])
        data_rows = [
            dict(zip(headers, r)) for r in rows[1:]
            if r and r[0] not in (None, '')   # skip blank + total row
        ]

        self.stdout.write(f'  • Filas detectadas: {len(data_rows)}')

        # Phase 1 — Brands
        self.stdout.write(self.style.SUCCESS('\n--- FASE 1: MARCAS ---'))
        marcas_unicas = sorted({(r.get('Marca') or '').strip().upper() for r in data_rows if r.get('Marca')})
        for nombre in marcas_unicas:
            if not nombre:
                continue
            _, created = Brand.objects.update_or_create(name=nombre, defaults={})
            tag = '[+]' if created else '[-]'
            self.stdout.write(f'  {tag} {nombre}')

        # Phase 2 — TipoProducto + CampoProducto
        self.stdout.write(self.style.SUCCESS('\n--- FASE 2: TIPO PRODUCTO Y CAMPOS ---'))
        tipo_laptop, _ = TipoProducto.objects.update_or_create(nombre='Laptop', defaults={})
        for orden, nombre_campo in enumerate(LAPTOP_FIELDS, start=1):
            campo, _ = CampoProducto.objects.update_or_create(
                nombre=nombre_campo,
                defaults={'tipo': 'texto'},
            )
            TipoProductoCampo.objects.update_or_create(
                tipo_producto=tipo_laptop,
                campo_producto=campo,
                defaults={'required': True, 'orden': orden},
            )
            self.stdout.write(f'  • {nombre_campo}')

        # Phase 3 — Proveedor default
        self.stdout.write(self.style.SUCCESS('\n--- FASE 3: PROVEEDOR ---'))
        proveedor, created = Proveedor.objects.update_or_create(
            nombre=PROVEEDOR_DEFAULT, defaults={},
        )
        self.stdout.write(f'  {"[+]" if created else "[-]"} {PROVEEDOR_DEFAULT}')

        # Phase 4 — Productos + Phase 5 — Órdenes / Unidades
        self.stdout.write(self.style.SUCCESS('\n--- FASE 4-5: PRODUCTOS, ORDENES Y UNIDADES ---'))
        campos_map = {c.nombre: c for c in CampoProducto.objects.filter(nombre__in=LAPTOP_FIELDS)}

        n_prod = 0
        n_ord = 0
        for idx, row in enumerate(data_rows):
            nombre = (row.get('Nombre') or '').strip()
            descripcion = (row.get('Descripción') or '').strip()
            marca_name = (row.get('Marca') or '').strip().upper()
            estado_excel = (row.get('Estado') or '').strip().lower()
            entrega_excel = (row.get('Entrega') or '').strip().lower()
            precio_cop_raw = row.get('Precio (COP)')

            if not (nombre and marca_name and estado_excel and entrega_excel and precio_cop_raw):
                self.stdout.write(self.style.WARNING(f'  [-] Fila {idx+1}: incompleta, saltando'))
                continue

            condicion = ESTADO_TO_CONDICION.get(estado_excel)
            estado_logistico = ENTREGA_TO_ESTADO_LOGISTICO.get(entrega_excel)
            if not condicion or not estado_logistico:
                self.stdout.write(self.style.WARNING(
                    f'  [-] Fila {idx+1}: estado/entrega no mapeable ({estado_excel}/{entrega_excel})'
                ))
                continue

            try:
                marca = Brand.objects.get(name__iexact=marca_name)
            except Brand.DoesNotExist:
                self.stdout.write(self.style.ERROR(f'  [!] Marca {marca_name} no existe'))
                continue

            # Producto upsert
            producto, prod_created = Producto.objects.update_or_create(
                nombre=nombre,
                defaults={
                    'descripcion': descripcion or nombre,
                    'marca': marca,
                    'tipo_producto': tipo_laptop,
                    'usuario_ultima_modificacion': usuario,
                },
            )

            # Dynamic field values (texto)
            for campo_nombre, campo in campos_map.items():
                excel_value = row.get(campo_nombre)
                if excel_value is None or str(excel_value).strip() == '':
                    continue
                ProductoCampoValor.objects.update_or_create(
                    producto=producto,
                    campo_producto=campo,
                    defaults={'valor_texto': str(excel_value).strip()},
                )

            n_prod += 1 if prod_created else 0

            # OrdenCompra — only create if there isn't already one for this producto+condicion.
            # Uniqueness is by numero_orden; we generate a stable one from the row index.
            numero_orden = f'LAP-INV-{idx+1:03d}'
            existing = OrdenCompra.objects.filter(numero_orden=numero_orden).first()
            if existing:
                self.stdout.write(f'  [-] Orden {numero_orden} ya existía ({nombre})')
                continue

            precio_cop = Decimal(str(precio_cop_raw))
            costo_compra_usd = (precio_cop / DEFAULT_TRM_COP / ESTIMATED_MARGIN).quantize(Decimal('0.01'))
            fecha_compra = FECHAS_COMPRA[idx % len(FECHAS_COMPRA)]

            orden = OrdenCompra(
                producto=producto,
                condicion=condicion,
                estado_logistico=estado_logistico,
                proveedor=proveedor,
                numero_orden=numero_orden,
                costo_compra=costo_compra_usd,
                fecha_compra=fecha_compra,
                usuario_ultima_modificacion=usuario,
            )
            orden._pct_impuesto = Decimal('2')
            orden.save()

            # Override unit price to match Excel COP price (the auto-calc in
            # OrdenCompra.save uses costo_compra * 1.2 * TRM which is just an
            # approximation; the real selling price is what the user listed).
            unidad = orden.unidad_producto
            if unidad:
                unidad.precio = precio_cop
                unidad.save(update_fields=['precio'])

            n_ord += 1
            self.stdout.write(self.style.SUCCESS(
                f'  [+] {numero_orden}: {nombre} — {condicion}/{estado_logistico} — '
                f'${precio_cop:,.0f} COP — fecha_compra {fecha_compra}'
            ))

        self.stdout.write(self.style.SUCCESS(
            f'\n=== HECHO: productos nuevos={n_prod}, ordenes nuevas={n_ord} ==='
        ))

    def _get_superuser(self):
        usuario = User.objects.filter(is_superuser=True).first()
        if not usuario:
            self.stdout.write(self.style.ERROR(
                'No superuser found. Create one with: python manage.py createsuperuser'
            ))
        return usuario
