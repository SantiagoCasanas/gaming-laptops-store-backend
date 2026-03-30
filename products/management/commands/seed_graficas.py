from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from products.models import (Brand, TipoProducto, CampoProducto, TipoProductoCampo,
                             Proveedor, Producto, ProductoCampoValor, BajoPedido,
                             UnidadProducto)
from purchases.models import OrdenCompra
from sales.models import Cliente
from decimal import Decimal
import re

User = get_user_model()


class Command(BaseCommand):
    help = 'Seed all GPU/Video Card data: brands, product types, supplier types, and dynamic fields'

    def handle(self, *args, **options):
        usuario = self._get_superuser()
        if not usuario:
            return

        self.stdout.write(self.style.SUCCESS('\n=== INICIANDO CARGA DE DATOS - TARJETAS DE VIDEO ===\n'))

        # Fase 1: Marcas
        self.stdout.write(self.style.SUCCESS('\n--- FASE 1: MARCAS ---'))
        self._seed_brands(usuario)

        # Fase 3: Tipo de Producto + Campos Dinámicos
        self.stdout.write(self.style.SUCCESS('\n--- FASE 3: TIPO DE PRODUCTO + CAMPOS DINAMICOS ---'))
        self._seed_tipo_grafica(usuario)

        # Fase 4: Proveedores
        self.stdout.write(self.style.SUCCESS('\n--- FASE 4: PROVEEDORES ---'))
        self._seed_providers(usuario)

        # Fase 5: Productos GPU
        self.stdout.write(self.style.SUCCESS('\n--- FASE 5: PRODUCTOS GPU ---'))
        self._seed_productos(usuario)

        # Fase 6: BajoPedido (On-Demand Sourcing)
        self.stdout.write(self.style.SUCCESS('\n--- FASE 6: BAJO PEDIDO (SOURCING BAJO DEMANDA) ---'))
        self._seed_variantes(usuario)

        # Fase 7: Unidades
        self.stdout.write(self.style.SUCCESS('\n--- FASE 7: UNIDADES ---'))
        self._seed_unidades(usuario)

        # Fase 8: Órdenes de Compra
        self.stdout.write(self.style.SUCCESS('\n--- FASE 8: ORDENES DE COMPRA ---'))
        self._seed_ordenes_compra(usuario)

        self.stdout.write(self.style.SUCCESS('\n=== CARGA COMPLETADA ===\n'))

    def _get_superuser(self):
        """Obtiene el primer superuser o muestra error."""
        try:
            usuario = User.objects.filter(is_superuser=True).first()
            if not usuario:
                self.stdout.write(self.style.ERROR('No superuser found. Create one with: python manage.py createsuperuser'))
                return None
            return usuario
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error getting superuser: {str(e)}'))
            return None

    def _seed_brands(self, usuario):
        """Fase 1: Cargar marcas desde Excel."""
        try:
            wb = load_workbook('products/datos-precarga/graficas/archivos_de_carga/01_Fase1_Marcas.xlsx')
            ws = wb.active
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR('Archivo no encontrado: products/datos-precarga/graficas/archivos_de_carga/01_Fase1_Marcas.xlsx'))
            return

        brands = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value]
        created_count = 0
        existing_count = 0

        for nombre in brands:
            nombre = str(nombre).strip() if nombre else ''
            if not nombre:
                self.stdout.write(self.style.WARNING('Marca sin nombre, saltando...'))
                continue

            try:
                brand, created = Brand.objects.update_or_create(
                    name=nombre,
                    defaults={}
                )

                if created:
                    self.stdout.write(self.style.SUCCESS(f'[+] Creado: {nombre}'))
                    created_count += 1
                else:
                    self.stdout.write(self.style.WARNING(f'[-] Ya existía: {nombre}'))
                    existing_count += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'[!] Error creando "{nombre}": {str(e)}'))

        self.stdout.write(self.style.SUCCESS(f'\nResumen Fase 1: {created_count} nuevas, {existing_count} existentes, {len(brands)} total'))

    def _seed_tipo_grafica(self, usuario):
        """Fase 3: Crear TipoProducto y CampoProducto (hardcoded)."""
        campos_definicion = [
            {'nombre': 'Fabricante del Chip', 'tipo': 'texto', 'required': True},
            {'nombre': 'Chip GPU', 'tipo': 'texto', 'required': True},
            {'nombre': 'Stream Processors / CUDA Cores', 'tipo': 'numero', 'required': True},
            {'nombre': 'Velocidad Boost (MHz)', 'tipo': 'numero', 'required': True},
            {'nombre': 'VRAM (GB)', 'tipo': 'numero', 'required': True},
            {'nombre': 'Tipo de Memoria', 'tipo': 'texto', 'required': True},
            {'nombre': 'Ancho de Bus (bits)', 'tipo': 'numero', 'required': True},
            {'nombre': 'Consumo de Energía (W)', 'tipo': 'numero', 'required': True},
            {'nombre': 'Fuente Recomendada (W)', 'tipo': 'numero', 'required': True},
            {'nombre': 'Conector de Poder', 'tipo': 'texto', 'required': True},
            {'nombre': 'Puertos de Video', 'tipo': 'texto', 'required': True},
            {'nombre': 'Interfaz PCIe', 'tipo': 'texto', 'required': True},
            {'nombre': 'Refrigeración', 'tipo': 'texto', 'required': True},
            {'nombre': 'Longitud (mm)', 'tipo': 'numero', 'required': True},
        ]

        # Crear TipoProducto
        try:
            tipo_producto, tipo_created = TipoProducto.objects.update_or_create(
                nombre='Tarjeta de Video',
                defaults={}
            )
            if tipo_created:
                self.stdout.write(self.style.SUCCESS('[+] Creado: TipoProducto "Tarjeta de Video"'))
            else:
                self.stdout.write(self.style.WARNING('[-] Ya existía: TipoProducto "Tarjeta de Video"'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'[!] Error creando TipoProducto: {str(e)}'))
            return

        # Crear campos dinámicos
        created_campos = 0
        existing_campos = 0
        orden = 1

        for campo_def in campos_definicion:
            try:
                campo_producto, campo_created = CampoProducto.objects.update_or_create(
                    nombre=campo_def['nombre'],
                    defaults={'tipo': campo_def['tipo']}
                )

                tipo_campo, tipo_campo_created = TipoProductoCampo.objects.update_or_create(
                    tipo_producto=tipo_producto,
                    campo_producto=campo_producto,
                    defaults={
                        'required': campo_def['required'],
                        'orden': orden,
                    }
                )

                if campo_created or tipo_campo_created:
                    self.stdout.write(self.style.SUCCESS(f'[+] {campo_def["nombre"]}'))
                    created_campos += 1
                else:
                    self.stdout.write(self.style.WARNING(f'[-] {campo_def["nombre"]}'))
                    existing_campos += 1

                orden += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'[!] Error en "{campo_def["nombre"]}": {str(e)}'))

        self.stdout.write(self.style.SUCCESS(f'\nResumen Fase 3: {created_campos} nuevos, {existing_campos} existentes, {len(campos_definicion)} total'))

    def _seed_providers(self, usuario):
        """Fase 4: Cargar proveedores desde Excel."""
        try:
            wb = load_workbook('products/datos-precarga/graficas/archivos_de_carga/04_Fase4_Proveedores.xlsx')
            ws = wb.active
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR('Archivo no encontrado: products/datos-precarga/graficas/archivos_de_carga/04_Fase4_Proveedores.xlsx'))
            return

        proveedores = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value]
        created_count = 0
        existing_count = 0

        for nombre in proveedores:
            nombre = str(nombre).strip() if nombre else ''
            if not nombre:
                self.stdout.write(self.style.WARNING('Proveedor sin nombre, saltando...'))
                continue

            try:
                proveedor, created = Proveedor.objects.update_or_create(
                    nombre=nombre,
                    defaults={}
                )

                if created:
                    self.stdout.write(self.style.SUCCESS(f'[+] Creado: {nombre}'))
                    created_count += 1
                else:
                    self.stdout.write(self.style.WARNING(f'[-] Ya existía: {nombre}'))
                    existing_count += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'[!] Error creando "{nombre}": {str(e)}'))

        self.stdout.write(self.style.SUCCESS(f'\nResumen Fase 4: {created_count} nuevos, {existing_count} existentes, {len(proveedores)} total'))

    def _seed_productos(self, usuario):
        """Fase 5: Cargar productos GPU desde Excel."""
        try:
            wb = load_workbook('products/datos-precarga/graficas/archivos_de_carga/05_Fase5_Productos.xlsx')
            ws = wb.active
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR('Archivo no encontrado: products/datos-precarga/graficas/archivos_de_carga/05_Fase5_Productos.xlsx'))
            return

        # Obtener referencias
        try:
            tipo_grafica = TipoProducto.objects.get(nombre='Tarjeta de Video')
        except TipoProducto.DoesNotExist:
            self.stdout.write(self.style.ERROR('[!] TipoProducto "Tarjeta de Video" no existe. Ejecutar Fase 3 primero.'))
            return

        # Mapeo de nombres en Excel a nombres en BD
        excel_to_bd_mapping = {
            'fabricante_chip': 'Fabricante del Chip',
            'chip_gpu': 'Chip GPU',
            'stream_processors': 'Stream Processors / CUDA Cores',
            'velocidad_boost_mhz': 'Velocidad Boost (MHz)',
            'vram_gb': 'VRAM (GB)',
            'tipo_memoria': 'Tipo de Memoria',
            'ancho_bus_bits': 'Ancho de Bus (bits)',
            'consumo_w': 'Consumo de Energía (W)',
            'fuente_recomendada_w': 'Fuente Recomendada (W)',
            'conector_poder': 'Conector de Poder',
            'puertos_video': 'Puertos de Video',
            'interfaz_pcie': 'Interfaz PCIe',
            'refrigeracion': 'Refrigeración',
            'longitud_mm': 'Longitud (mm)',
        }

        # Obtener campos dinámicos por nombre (como está en BD)
        campos_mapa = {}
        for campo in CampoProducto.objects.all():
            campos_mapa[campo.nombre] = campo

        # Leer Excel
        headers = [cell.value for cell in ws[1]]
        productos_data = list(ws.iter_rows(min_row=2, values_only=True))

        created_count = 0
        existing_count = 0

        for producto_row in productos_data:
            # Convertir a diccionario
            producto_dict = {}
            for i, header in enumerate(headers):
                if i < len(producto_row):
                    producto_dict[header] = producto_row[i]

            nombre = str(producto_dict.get('nombre', '')).strip() if producto_dict.get('nombre') else ''
            if not nombre:
                self.stdout.write(self.style.WARNING('Producto sin nombre, saltando...'))
                continue

            try:
                # Obtener marca (case-insensitive)
                marca_nombre = producto_dict.get('marca', '').strip()
                try:
                    marca = Brand.objects.get(name__iexact=marca_nombre)
                except Brand.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'[!] Marca "{marca_nombre}" no existe en BD. Saltando producto "{nombre}"'))
                    continue

                # Crear o actualizar producto
                producto, created = Producto.objects.update_or_create(
                    nombre=nombre,
                    defaults={
                        'descripcion': producto_dict.get('descripcion', ''),
                        'marca': marca,
                        'tipo_producto': tipo_grafica,
                        'usuario_ultima_modificacion': usuario
                    }
                )

                # Cargar campos dinámicos
                for excel_col, valor in producto_dict.items():
                    if excel_col not in ['nombre', 'descripcion', 'marca']:
                        # Mapear nombre de Excel a nombre en BD
                        nombre_campo_bd = excel_to_bd_mapping.get(excel_col)
                        if nombre_campo_bd and nombre_campo_bd in campos_mapa:
                            campo = campos_mapa[nombre_campo_bd]

                            if valor is not None and str(valor).strip():
                                valor_str = str(valor).strip()

                                # Determinar tipo de valor según el campo
                                try:
                                    if campo.tipo == 'numero':
                                        valor_numero = float(valor_str) if valor_str else None
                                        ProductoCampoValor.objects.update_or_create(
                                            producto=producto,
                                            campo_producto=campo,
                                            defaults={'valor_numero': valor_numero}
                                        )
                                    elif campo.tipo == 'booleano':
                                        valor_bool = valor_str.lower() in ['true', '1', 'si', 'yes', 'verdadero']
                                        ProductoCampoValor.objects.update_or_create(
                                            producto=producto,
                                            campo_producto=campo,
                                            defaults={'valor_booleano': valor_bool}
                                        )
                                    else:  # texto
                                        ProductoCampoValor.objects.update_or_create(
                                            producto=producto,
                                            campo_producto=campo,
                                            defaults={'valor_texto': valor_str}
                                        )
                                except Exception as e:
                                    self.stdout.write(self.style.ERROR(f'[!] Error en campo "{excel_col}" (BD: {nombre_campo_bd}) de "{nombre}": {str(e)}'))

                if created:
                    self.stdout.write(self.style.SUCCESS(f'[+] Creado: {nombre}'))
                    created_count += 1
                else:
                    self.stdout.write(self.style.WARNING(f'[-] Ya existía: {nombre}'))
                    existing_count += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'[!] Error procesando "{nombre}": {str(e)}'))

        self.stdout.write(self.style.SUCCESS(f'\nResumen Fase 5: {created_count} nuevos, {existing_count} existentes, {len(productos_data)} total'))

    def _seed_variantes(self, usuario):
        """Fase 6: Cargar BajoPedido (sourcing bajo demanda) desde Excel."""
        try:
            wb = load_workbook('products/datos-precarga/graficas/archivos_de_carga/06_Fase6_Variantes.xlsx')
            ws = wb.active
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR('Archivo no encontrado: products/datos-precarga/graficas/archivos_de_carga/06_Fase6_Variantes.xlsx'))
            return

        # Leer Excel
        headers = [cell.value for cell in ws[1]]
        variantes_data = list(ws.iter_rows(min_row=2, values_only=True))

        created_count = 0
        existing_count = 0

        for variante_row in variantes_data:
            # Convertir a diccionario
            variante_dict = {}
            for i, header in enumerate(headers):
                if i < len(variante_row):
                    variante_dict[header] = variante_row[i]

            producto_nombre = variante_dict.get('producto_nombre', '').strip() if variante_dict.get('producto_nombre') else ''
            condicion = variante_dict.get('condicion', '').strip().lower() if variante_dict.get('condicion') else ''
            estado = variante_dict.get('estado', '').strip().lower() if variante_dict.get('estado') else ''
            precio_str = variante_dict.get('precio', '')

            if not producto_nombre or not condicion or not estado or not precio_str:
                self.stdout.write(self.style.WARNING('Variante incompleta (falta producto_nombre, condicion, estado o precio), saltando...'))
                continue

            try:
                # Obtener producto
                try:
                    producto = Producto.objects.get(nombre__iexact=producto_nombre)
                except Producto.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'[!] Producto "{producto_nombre}" no existe. Saltando variante.'))
                    continue

                # Parsear precio (remover $ y . de miles)
                try:
                    precio_limpio = re.sub(r'[\$\.]', '', str(precio_str)).replace(',', '.')
                    precio = Decimal(precio_limpio)
                except (ValueError, TypeError):
                    self.stdout.write(self.style.ERROR(f'[!] Precio inválido "{precio_str}" para {producto_nombre}. Saltando.'))
                    continue

                # Validar enumerados
                condiciones_validas = [c[0] for c in BajoPedido.CondicionChoices.choices]
                estados_validos = [e[0] for e in BajoPedido.EstadoChoices.choices]

                if condicion not in condiciones_validas:
                    self.stdout.write(self.style.ERROR(f'[!] Condición inválida "{condicion}". Debe ser: {", ".join(condiciones_validas)}'))
                    continue

                if estado not in estados_validos:
                    self.stdout.write(self.style.ERROR(f'[!] Estado inválido "{estado}". Debe ser: {", ".join(estados_validos)}'))
                    continue

                # Obtener proveedor (opcional, case-insensitive)
                proveedor = None
                proveedor_nombre = variante_dict.get('proveedor', '').strip() if variante_dict.get('proveedor') else ''
                if proveedor_nombre:
                    try:
                        proveedor = Proveedor.objects.get(nombre__iexact=proveedor_nombre)
                    except Proveedor.DoesNotExist:
                        self.stdout.write(self.style.WARNING(f'[-] Proveedor "{proveedor_nombre}" no existe, se ignorará'))
                        proveedor = None

                # Crear o actualizar variante
                enlace = variante_dict.get('enlace_proveedor', '').strip() if variante_dict.get('enlace_proveedor') else None

                # Usar unique_together de producto + condicion + estado para detectar si ya existe
                variante, created = BajoPedido.objects.update_or_create(
                    producto=producto,
                    condicion=condicion,
                    estado=estado,
                    defaults={
                        'precio': precio,
                        'proveedor': proveedor,
                        'enlace_proveedor': enlace,
                        'usuario_ultima_modificacion': usuario
                    }
                )

                if created:
                    self.stdout.write(self.style.SUCCESS(f'[+] Creado: {producto_nombre} ({condicion}/{estado}) - ${precio:,.0f}'))
                    created_count += 1
                else:
                    self.stdout.write(self.style.WARNING(f'[-] Ya existía: {producto_nombre} ({condicion}/{estado})'))
                    existing_count += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'[!] Error procesando variante de "{producto_nombre}": {str(e)}'))

        self.stdout.write(self.style.SUCCESS(f'\nResumen Fase 6: {created_count} nuevas, {existing_count} existentes, {len(variantes_data)} total'))

    def _seed_unidades(self, usuario):
        """Fase 7: Cargar unidades físicas desde Excel."""
        try:
            wb = load_workbook('products/datos-precarga/graficas/archivos_de_carga/07_Fase7_Unidades.xlsx')
            ws = wb.active
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR('Archivo no encontrado: products/datos-precarga/graficas/archivos_de_carga/07_Fase7_Unidades.xlsx'))
            return

        # Leer Excel
        headers = [cell.value for cell in ws[1]]
        unidades_data = list(ws.iter_rows(min_row=2, values_only=True))

        created_count = 0
        existing_count = 0

        for unidad_row in unidades_data:
            # Convertir a diccionario
            unidad_dict = {}
            for i, header in enumerate(headers):
                if i < len(unidad_row):
                    unidad_dict[header] = unidad_row[i]

            producto_nombre = unidad_dict.get('producto_nombre', '').strip() if unidad_dict.get('producto_nombre') else ''
            condicion = unidad_dict.get('condicion', '').strip().lower() if unidad_dict.get('condicion') else ''
            serial = unidad_dict.get('serial', '').strip() if unidad_dict.get('serial') else ''
            estado_venta = unidad_dict.get('estado_venta', '').strip().lower() if unidad_dict.get('estado_venta') else ''
            estado_producto = unidad_dict.get('estado_producto', '').strip().lower() if unidad_dict.get('estado_producto') else ''
            precio_str = unidad_dict.get('precio', '')

            if not producto_nombre or not condicion or not serial or not estado_venta or not estado_producto or not precio_str:
                self.stdout.write(self.style.WARNING('Unidad incompleta (falta alguno de: producto_nombre, condicion, serial, estado_venta, estado_producto o precio), saltando...'))
                continue

            try:
                # Obtener producto
                try:
                    producto = Producto.objects.get(nombre__iexact=producto_nombre)
                except Producto.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'[!] Producto "{producto_nombre}" no existe. Saltando unidad {serial}.'))
                    continue

                # Obtener variante por producto + condicion
                try:
                    variante = BajoPedido.objects.get(
                        producto=producto,
                        condicion=condicion
                    )
                except BajoPedido.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'[!] Variante de "{producto_nombre}" ({condicion}) no existe. Saltando unidad {serial}.'))
                    continue
                except BajoPedido.MultipleObjectsReturned:
                    self.stdout.write(self.style.ERROR(f'[!] Múltiples variantes para "{producto_nombre}" ({condicion}). Saltando unidad {serial}.'))
                    continue

                # Parsear precio
                try:
                    precio_limpio = re.sub(r'[\$\.]', '', str(precio_str)).replace(',', '.')
                    precio = Decimal(precio_limpio)
                except (ValueError, TypeError):
                    self.stdout.write(self.style.ERROR(f'[!] Precio inválido "{precio_str}" para unidad {serial}. Saltando.'))
                    continue

                # Validar enumerados
                estados_venta_validos = [e[0] for e in UnidadProducto.EstadoVentaChoices.choices]
                estados_producto_validos = [e[0] for e in UnidadProducto.EstadoProductoChoices.choices]

                if estado_venta not in estados_venta_validos:
                    self.stdout.write(self.style.ERROR(f'[!] Estado de venta inválido "{estado_venta}" en serial {serial}. Válidos: {", ".join(estados_venta_validos)}'))
                    continue

                if estado_producto not in estados_producto_validos:
                    self.stdout.write(self.style.ERROR(f'[!] Estado de producto inválido "{estado_producto}" en serial {serial}. Válidos: {", ".join(estados_producto_validos)}'))
                    continue

                # Crear o actualizar unidad
                unidad, created = UnidadProducto.objects.update_or_create(
                    serial=serial,
                    defaults={
                        'producto': producto,
                        'condicion': condicion,
                        'estado_venta': estado_venta,
                        'estado_producto': estado_producto,
                        'precio': precio,
                        'usuario_ultima_modificacion': usuario
                    }
                )

                if created:
                    self.stdout.write(self.style.SUCCESS(f'[+] Creado: {serial} ({producto_nombre}) - ${precio:,.0f}'))
                    created_count += 1
                else:
                    self.stdout.write(self.style.WARNING(f'[-] Ya existía: {serial}'))
                    existing_count += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'[!] Error procesando unidad {serial}: {str(e)}'))

        self.stdout.write(self.style.SUCCESS(f'\nResumen Fase 7: {created_count} nuevas, {existing_count} existentes, {len(unidades_data)} total'))

    def _seed_ordenes_compra(self, usuario):
        """Fase 8: Cargar órdenes de compra desde Excel."""
        try:
            wb = load_workbook('products/datos-precarga/graficas/archivos_de_carga/08_Fase8_OrdenesCompra.xlsx')
            ws = wb.active
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR('Archivo no encontrado: products/datos-precarga/graficas/archivos_de_carga/08_Fase8_OrdenesCompra.xlsx'))
            return

        # Leer Excel
        headers = [cell.value for cell in ws[1]]
        ordenes_data = list(ws.iter_rows(min_row=2, values_only=True))

        created_count = 0
        existing_count = 0

        for orden_row in ordenes_data:
            # Convertir a diccionario
            orden_dict = {}
            for i, header in enumerate(headers):
                if i < len(orden_row):
                    orden_dict[header] = orden_row[i]

            producto_nombre = orden_dict.get('producto_nombre', '').strip() if orden_dict.get('producto_nombre') else ''
            condicion = orden_dict.get('condicion', '').strip().lower() if orden_dict.get('condicion') else ''
            numero_orden = orden_dict.get('numero_orden', '').strip() if orden_dict.get('numero_orden') else ''
            tipo = orden_dict.get('tipo', '').strip().lower() if orden_dict.get('tipo') else ''
            estado_logistico = orden_dict.get('estado_logistico', '').strip().lower() if orden_dict.get('estado_logistico') else ''
            costo_compra_str = orden_dict.get('costo_compra', '')

            if not producto_nombre or not condicion or not numero_orden or not tipo or not estado_logistico or not costo_compra_str:
                self.stdout.write(self.style.WARNING('Orden incompleta (falta alguno de: producto, condicion, numero_orden, tipo, estado_logistico o costo_compra), saltando...'))
                continue

            try:
                # Obtener producto y variante
                try:
                    producto = Producto.objects.get(nombre__iexact=producto_nombre)
                except Producto.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'[!] Producto "{producto_nombre}" no existe. Saltando orden {numero_orden}.'))
                    continue

                try:
                    variante = BajoPedido.objects.get(producto=producto, condicion=condicion)
                except BajoPedido.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'[!] Variante de "{producto_nombre}" ({condicion}) no existe. Saltando orden {numero_orden}.'))
                    continue

                # Parsear costo_compra
                try:
                    costo_compra = Decimal(str(costo_compra_str))
                except (ValueError, TypeError):
                    self.stdout.write(self.style.ERROR(f'[!] Costo de compra inválido "{costo_compra_str}". Saltando orden {numero_orden}.'))
                    continue

                # Validar tipo
                tipos_validos = [t[0] for t in OrdenCompra.TipoChoices.choices]
                if tipo not in tipos_validos:
                    self.stdout.write(self.style.ERROR(f'[!] Tipo inválido "{tipo}". Debe ser: {", ".join(tipos_validos)}'))
                    continue

                # Validar estado_logistico
                estados_validos = [e[0] for e in OrdenCompra.EstadoLogisticoChoices.choices]
                if estado_logistico not in estados_validos:
                    self.stdout.write(self.style.ERROR(f'[!] Estado logístico inválido "{estado_logistico}". Debe ser: {", ".join(estados_validos)}'))
                    continue

                # Obtener proveedor (opcional)
                proveedor = None
                proveedor_nombre = orden_dict.get('proveedor', '').strip() if orden_dict.get('proveedor') else ''
                if proveedor_nombre:
                    try:
                        proveedor = Proveedor.objects.get(nombre__iexact=proveedor_nombre)
                    except Proveedor.DoesNotExist:
                        self.stdout.write(self.style.WARNING(f'[-] Proveedor "{proveedor_nombre}" no existe, se ignorará'))

                # Obtener cliente (opcional, solo si tipo = canje_cliente)
                cliente = None
                if tipo == 'canje_cliente':
                    cliente_cedula = orden_dict.get('cliente_cedula', '').strip() if orden_dict.get('cliente_cedula') else ''
                    if cliente_cedula:
                        try:
                            cliente = Cliente.objects.get(cedula=cliente_cedula)
                        except Cliente.DoesNotExist:
                            self.stdout.write(self.style.ERROR(f'[!] Cliente con cédula "{cliente_cedula}" no existe. Saltando orden {numero_orden}.'))
                            continue
                    else:
                        self.stdout.write(self.style.ERROR(f'[!] Tipo canje_cliente requiere cliente_cedula. Saltando orden {numero_orden}.'))
                        continue

                # Parsear costo_importacion (opcional)
                costo_importacion = None
                costo_importacion_str = orden_dict.get('costo_importacion', '')
                if costo_importacion_str:
                    try:
                        costo_importacion = Decimal(str(costo_importacion_str))
                    except (ValueError, TypeError):
                        self.stdout.write(self.style.WARNING(f'[-] Costo de importación inválido, se ignorará'))

                # Datos opcionales
                numero_tracking = orden_dict.get('numero_tracking', '').strip() if orden_dict.get('numero_tracking') else None
                serial_generado = orden_dict.get('serial_generado', '').strip() if orden_dict.get('serial_generado') else None

                # Crear o actualizar orden
                orden, created = OrdenCompra.objects.update_or_create(
                    numero_orden=numero_orden,
                    defaults={
                        'producto': producto,
                        'condicion': condicion,
                        'tipo': tipo,
                        'estado_logistico': estado_logistico,
                        'proveedor': proveedor,
                        'cliente': cliente,
                        'numero_tracking': numero_tracking,
                        'costo_compra': costo_compra,
                        'costo_importacion': costo_importacion,
                        'serial_generado': serial_generado,
                        'usuario_ultima_modificacion': usuario
                    }
                )

                # La unidad se crea automáticamente en el método save() de OrdenCompra
                unidad_serial = orden.unidad_producto.serial if orden.unidad_producto else "pending"

                if created:
                    self.stdout.write(self.style.SUCCESS(f'[+] Creado: {numero_orden} - {producto_nombre} (Unit: {unidad_serial})'))
                    created_count += 1
                else:
                    self.stdout.write(self.style.WARNING(f'[-] Ya existía: {numero_orden}'))
                    existing_count += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'[!] Error procesando orden {numero_orden}: {str(e)}'))

        self.stdout.write(self.style.SUCCESS(f'\nResumen Fase 8: {created_count} nuevas, {existing_count} existentes, {len(ordenes_data)} total'))
