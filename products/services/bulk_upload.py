"""
Bulk product upload from Excel templates.

Public functions:
  • build_template_workbook(tipo_producto)
        Returns an openpyxl Workbook with the dynamic columns required to load
        products of the given TipoProducto. Row 1 = header (column name),
        row 2 = type hint ("(texto)", "(numero)", "(booleano)" + obligatorio
        marker). Data starts at row 3.

  • parse_and_validate(tipo_producto, file, dry_run, usuario)
        Reads the uploaded xlsx and validates each row. Returns
            {
                "dry_run": bool,
                "total": int,
                "creados":      [{"fila": n, "nombre": ..., "data": {...}}, ...],
                "actualizados": [{"fila": n, "nombre": ..., "data": {...}}, ...],
                "fallidos":     [{"fila": n, "nombre": ..., "data": {...},
                                  "errores": [str, ...]}, ...],
            }
        Idempotency by Producto.nombre (case-insensitive): rows whose nombre
        matches an existing Producto go to `actualizados` and, when
        dry_run=False, the existing record (descripcion, marca,
        ProductoCampoValor) is overwritten with the row's values. Brand-new
        rows go to `creados`.

  • confirm_edited_rows(tipo_producto, rows, usuario)
        Re-validates and persists rows previously returned in the preview and
        edited by the user on the frontend. `rows` is a list of dicts with at
        minimum a `data` key containing the column → value map. Returns
            {
                "creados":      [{"id": int, "nombre": str, "marca_nombre": str,
                                  "imagenes_count": int}, ...],
                "actualizados": [{"id": int, "nombre": str, "marca_nombre": str,
                                  "imagenes_count": int}, ...],
                "fallidos":     [{"nombre": str, "errores": [...]}, ...],
            }
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from products.models import (
    Brand,
    CampoProducto,
    Producto,
    ProductoCampoValor,
    TipoProducto,
    TipoProductoCampo,
)

FIXED_COLUMNS = [
    ('nombre', 'texto', True),
    ('descripcion', 'texto', False),
    ('marca', 'texto', True),
]

TRUE_VALUES = {'true', 'verdadero', 'si', 'sí', 'yes', 'y', '1'}
FALSE_VALUES = {'false', 'falso', 'no', 'n', '0'}


def _ordered_dynamic_fields(tipo_producto: TipoProducto):
    """Return the TipoProductoCampo rows for this tipo, ordered by orden."""
    return list(
        TipoProductoCampo.objects
        .filter(tipo_producto=tipo_producto)
        .select_related('campo_producto')
        .order_by('orden', 'campo_producto__nombre')
    )


def build_template_workbook(tipo_producto: TipoProducto) -> Workbook:
    """Generate an Excel template for the given TipoProducto."""
    wb = Workbook()
    ws = wb.active
    ws.title = (tipo_producto.nombre or 'Plantilla')[:31]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2979C8')
    hint_font = Font(italic=True, color='6B7280', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    columns: list[tuple[str, str, bool]] = list(FIXED_COLUMNS)
    for tpc in _ordered_dynamic_fields(tipo_producto):
        columns.append((tpc.campo_producto.nombre, tpc.campo_producto.tipo, tpc.required))

    for col_idx, (header, tipo, required) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

        hint = f'({tipo}{" · obligatorio" if required else ""})'
        hint_cell = ws.cell(row=2, column=col_idx, value=hint)
        hint_cell.font = hint_font
        hint_cell.alignment = center

        ws.column_dimensions[cell.column_letter].width = max(16, len(header) + 4)

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = 'A3'
    return wb


# ---------------------------------------------------------------------------
# Parsing & validation
# ---------------------------------------------------------------------------


def _coerce_value(raw, tipo: str):
    """Return (parsed_value, error_or_None). raw=None or empty -> (None, None)."""
    if raw is None:
        return None, None
    if isinstance(raw, str):
        raw = raw.strip()
        if raw == '':
            return None, None

    if tipo == 'texto':
        return str(raw), None

    if tipo == 'numero':
        if isinstance(raw, (int, float, Decimal)):
            return Decimal(str(raw)), None
        try:
            return Decimal(str(raw).replace(',', '.')), None
        except (InvalidOperation, ValueError):
            return None, f'valor "{raw}" no es un número válido'

    if tipo == 'booleano':
        if isinstance(raw, bool):
            return raw, None
        s = str(raw).strip().lower()
        if s in TRUE_VALUES:
            return True, None
        if s in FALSE_VALUES:
            return False, None
        return None, f'valor "{raw}" no es booleano (use sí/no, true/false, 1/0)'

    return None, f'tipo de campo desconocido: {tipo}'


def _build_marca_index() -> dict[str, Brand]:
    """Lookup table keyed by lowercased brand name."""
    return {b.name.lower(): b for b in Brand.objects.all()}


def _apply_dynamic_values(producto, dynamic_values):
    """Recreate the ProductoCampoValor rows for `producto` from scratch."""
    ProductoCampoValor.objects.filter(producto=producto).delete()
    for campo, tipo, value in dynamic_values:
        kwargs = {'producto': producto, 'campo_producto': campo}
        if tipo == 'texto':
            kwargs['valor_texto'] = str(value)
        elif tipo == 'numero':
            kwargs['valor_numero'] = value
        elif tipo == 'booleano':
            kwargs['valor_booleano'] = bool(value)
        ProductoCampoValor.objects.create(**kwargs)


def _save_row(producto_data: dict, dynamic_values: list[tuple[CampoProducto, str, object]], usuario):
    """Persist a Producto + all its ProductoCampoValor rows atomically."""
    with transaction.atomic():
        producto = Producto.objects.create(
            nombre=producto_data['nombre'],
            descripcion=producto_data['descripcion'] or producto_data['nombre'],
            marca=producto_data['marca'],
            tipo_producto=producto_data['tipo_producto'],
            usuario_ultima_modificacion=usuario,
        )
        _apply_dynamic_values(producto, dynamic_values)
    return producto


def _update_row(producto, producto_data: dict, dynamic_values: list[tuple[CampoProducto, str, object]], usuario):
    """
    Overwrite an existing Producto with the values from a bulk-upload row.
    Updates descripcion, marca, tipo_producto and replaces the entire set of
    ProductoCampoValor rows. Nombre is the lookup key — never modified here.
    """
    with transaction.atomic():
        producto.descripcion = producto_data['descripcion'] or producto_data['nombre']
        producto.marca = producto_data['marca']
        producto.tipo_producto = producto_data['tipo_producto']
        producto.usuario_ultima_modificacion = usuario
        producto.save(update_fields=[
            'descripcion', 'marca', 'tipo_producto', 'usuario_ultima_modificacion',
        ])
        _apply_dynamic_values(producto, dynamic_values)
    return producto


def _serialize_row_data(nombre, descripcion, marca_name, dynamic_raw_by_name):
    """Build the `data` dict returned with every row in the preview."""
    out = {
        'nombre': nombre or '',
        'descripcion': descripcion or '',
        'marca': marca_name or '',
    }
    for k, v in dynamic_raw_by_name.items():
        # Always serialize as string; the editable preview is text-based.
        # Booleans become "true"/"false" so the frontend can show them in a
        # <select>; numbers become their str() form.
        if v is None:
            out[k] = ''
        elif isinstance(v, bool):
            out[k] = 'true' if v else 'false'
        else:
            out[k] = str(v)
    return out


def _validate_row(
    raw_data: dict,
    tipo_producto: TipoProducto,
    marca_index: dict,
    expected_dynamic: dict,
):
    """
    Validate a single row's `data` dict (column → raw value).
    Returns (status, nombre, payload) where status is one of:
        'valido'  → payload = {'producto_data': {...}, 'dynamic_values': [...]}
        'fallido' → payload = {'errores': [str, ...]}

    Whether a `valido` row results in an insert or an update is decided by the
    caller based on its own existing-product index — `_validate_row` does not
    look at duplicates.
    """
    nombre = str(raw_data.get('nombre') or '').strip()
    if not nombre:
        return 'fallido', '', {'errores': ['nombre es obligatorio']}

    errores: list[str] = []
    descripcion = str(raw_data.get('descripcion') or '').strip()

    marca_name = str(raw_data.get('marca') or '').strip()
    marca = marca_index.get(marca_name.lower()) if marca_name else None
    if not marca_name:
        errores.append('marca es obligatoria')
    elif marca is None:
        errores.append(f'la marca "{marca_name}" no existe en el sistema')

    dynamic_values: list[tuple[CampoProducto, str, object]] = []
    for nombre_campo, tpc in expected_dynamic.items():
        raw = raw_data.get(nombre_campo)
        value, err = _coerce_value(raw, tpc.campo_producto.tipo)
        if err:
            errores.append(f'{nombre_campo}: {err}')
            continue
        if value is None:
            if tpc.required:
                errores.append(f'{nombre_campo} es obligatorio')
            continue
        dynamic_values.append((tpc.campo_producto, tpc.campo_producto.tipo, value))

    if errores:
        return 'fallido', nombre, {'errores': errores}

    return 'valido', nombre, {
        'producto_data': {
            'nombre': nombre,
            'descripcion': descripcion,
            'marca': marca,
            'tipo_producto': tipo_producto,
        },
        'dynamic_values': dynamic_values,
    }


def parse_and_validate(tipo_producto: TipoProducto, uploaded_file, dry_run: bool, usuario) -> dict:
    """Main entry — reads the xlsx and returns the preview dict."""
    try:
        wb = load_workbook(BytesIO(uploaded_file.read()), data_only=True)
    except Exception as exc:
        return {
            'dry_run': dry_run,
            'total': 0,
            'creados': [],
            'actualizados': [],
            'fallidos': [{'fila': 0, 'nombre': '', 'data': {}, 'errores': [f'No se pudo leer el archivo: {exc}']}],
        }

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'dry_run': dry_run, 'total': 0, 'creados': [], 'actualizados': [], 'fallidos': []}

    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    header_index = {h: i for i, h in enumerate(headers) if h}

    dynamic_fields = _ordered_dynamic_fields(tipo_producto)
    expected_dynamic = {tpc.campo_producto.nombre: tpc for tpc in dynamic_fields}

    data_start = 1
    if len(rows) > 1:
        second_row = rows[1]
        if second_row and any(
            isinstance(v, str) and v.strip().startswith('(') for v in second_row if v is not None
        ):
            data_start = 2

    creados, actualizados, fallidos = [], [], []
    marca_index = _build_marca_index()
    # Map name (lowercased) → existing Producto, so we can decide insert vs.
    # update without re-querying on each row.
    existing_index = {p.nombre.lower(): p for p in Producto.objects.all()}

    for row_idx, raw_row in enumerate(rows[data_start:], start=data_start + 1):
        if raw_row is None or all(v in (None, '') for v in raw_row):
            continue

        def cell(col):
            i = header_index.get(col)
            if i is None or i >= len(raw_row):
                return None
            return raw_row[i]

        # Build the column → raw value map and the data preview snapshot
        nombre = str(cell('nombre') or '').strip()
        descripcion = str(cell('descripcion') or '').strip()
        marca_name = str(cell('marca') or '').strip()
        dynamic_raw_by_name = {n: cell(n) for n in expected_dynamic}
        data_preview = _serialize_row_data(nombre, descripcion, marca_name, dynamic_raw_by_name)

        raw_data_for_validation = {
            'nombre': nombre,
            'descripcion': descripcion,
            'marca': marca_name,
            **dynamic_raw_by_name,
        }
        status, returned_nombre, payload = _validate_row(
            raw_data_for_validation,
            tipo_producto,
            marca_index,
            expected_dynamic,
        )

        if status == 'fallido':
            fallidos.append({
                'fila': row_idx,
                'nombre': returned_nombre,
                'data': data_preview,
                'errores': payload['errores'],
            })
            continue

        producto_existente = existing_index.get(returned_nombre.lower())

        if not dry_run:
            try:
                if producto_existente:
                    _update_row(producto_existente, payload['producto_data'], payload['dynamic_values'], usuario)
                else:
                    nuevo = _save_row(payload['producto_data'], payload['dynamic_values'], usuario)
                    existing_index[returned_nombre.lower()] = nuevo
            except Exception as exc:
                fallidos.append({
                    'fila': row_idx,
                    'nombre': returned_nombre,
                    'data': data_preview,
                    'errores': [f'error al guardar: {exc}'],
                })
                continue

        bucket = actualizados if producto_existente else creados
        bucket.append({
            'fila': row_idx,
            'nombre': returned_nombre,
            'data': data_preview,
        })

    return {
        'dry_run': dry_run,
        'total': len(creados) + len(actualizados) + len(fallidos),
        'creados': creados,
        'actualizados': actualizados,
        'fallidos': fallidos,
    }


def confirm_edited_rows(tipo_producto: TipoProducto, rows: list, usuario) -> dict:
    """
    Persist edited rows coming from the JSON confirm endpoint. Each row should
    be `{ "data": {column: value, ...}, "fila": optional }`. Re-validates so
    edits cannot bypass the original checks. Rows whose `nombre` matches an
    existing Producto are overwritten in place; new names are inserted.
    """
    dynamic_fields = _ordered_dynamic_fields(tipo_producto)
    expected_dynamic = {tpc.campo_producto.nombre: tpc for tpc in dynamic_fields}
    marca_index = _build_marca_index()
    existing_index = {p.nombre.lower(): p for p in Producto.objects.all()}

    creados, actualizados, fallidos = [], [], []

    for row in rows or []:
        data = (row or {}).get('data') or {}
        fila = (row or {}).get('fila')

        status, nombre, payload = _validate_row(
            data, tipo_producto, marca_index, expected_dynamic
        )

        if status == 'fallido':
            fallidos.append({
                'fila': fila,
                'nombre': nombre,
                'data': data,
                'errores': payload['errores'],
            })
            continue

        producto_existente = existing_index.get(nombre.lower())

        try:
            if producto_existente:
                producto = _update_row(producto_existente, payload['producto_data'], payload['dynamic_values'], usuario)
                bucket = actualizados
            else:
                producto = _save_row(payload['producto_data'], payload['dynamic_values'], usuario)
                existing_index[nombre.lower()] = producto
                bucket = creados
        except Exception as exc:
            fallidos.append({
                'fila': fila,
                'nombre': nombre,
                'data': data,
                'errores': [f'error al guardar: {exc}'],
            })
            continue

        bucket.append({
            'fila': fila,
            'id': producto.id,
            'nombre': producto.nombre,
            'marca_nombre': producto.marca.name,
            'imagenes_count': producto.imagenes.count(),
        })

    return {
        'total': len(creados) + len(actualizados) + len(fallidos),
        'creados': creados,
        'actualizados': actualizados,
        'fallidos': fallidos,
    }
